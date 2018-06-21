import json
import requests
import csv
import secrets as secrets
from bs4 import BeautifulSoup
import pandas as pd
import re

from openpyxl import Workbook 

from openpyxl import load_workbook

import openpyxl

# api_key = secrets.api_key

FNAME = "Judeo_arabic_spreadsheet.xlsm"


## creating a class for each publication
class Publication(object):
	def __init__(self, title):
		self.title = title
		self.subject_tags = ""
	def __str__(self):
		return subject_tags
	pass

#opens excel file, writes subjects to the given subject_tag cell number --> WILL NEED TO CHANGE SPECIFIED CELL # LATER
def write_to_excel(write_to_cell_number, content_to_write):
	wb = Workbook()
	xfile = openpyxl.load_workbook(filename = FNAME, read_only=False, keep_vba=True)
	sheet = xfile.get_sheet_by_name('Judeo-Arabic')
	#print(sheet[write_to_cell_number].value)
	sheet[write_to_cell_number].value = content_to_write
	xfile.save(FNAME)


## perform reg ex on a term that is going to be used in the URL query
def reg_ex(query_term):
	improved_query_term = query_term.replace('[', ' ').replace(']', ' ').replace('.', ' ').replace(':', ' ').replace("ʻ", "").replace("(", "").replace(")", "").replace("ʼ", "").replace("'", "").lower().split()
	complete_query_string = ''
	for word in improved_query_term:
		complete_query_string += word + "-"
	complete_query_string = complete_query_string[:-1].replace("--", '-')

	return complete_query_string


#takes give title, eliminates uncessary characters and tokenizes, creates primary url. 
#if first word in title is "al" creates, secondary url
def assemble_url(a_title, an_author):
	title= reg_ex(a_given_title)
	baseurl = "https://www.worldcat.org/search?q=" + title

	## if there is an author:  need to get rid of the title part in the base url

	author = reg_ex(an_author)
	primary_url = baseurl + title + author
	return primary_url


##try to find subject div, if doesn't exist, meaning there are no subjects, return false and record the error in error_cell_number
##if there are subjects, return subject string. 
def find_subjects(soup, error_cell_number):
	#if the item has subject terms
	if soup.find_all(id = "subject-terms"):
		subject_div = soup.find_all(id = "subject-terms")
		# subject_div = subject_div.find_all('li', attrs={'class' :"subject-term"})
		subject_list = []
		subject_string = ""

		for line in subject_div:
			subjects = line.text.replace('--', "").replace(".", "").split()

			for item in subjects:
				if item not in subject_list:
					subject_list.append(item)
			subject_string = "; ".join(subject_list)
			#print(subject_string)
			return subject_string
	#the item doesn't have subject terms 
	else:
		##couldn't get subjects 
		print("Could not find subjects")
		error_message = "No subjects"
		write_to_excel(error_cell_number, error_message)
		return False			


## finds and returns description or false. does not record an error message
def find_page_number_description(soup):
	description_div = soup.find(id = "details-description")
	if description_div:
		description = description_div["td"].text
		print(description)
		return description
	else:
		print("no description")
		return False


## finds and returns notes or false. does not record an error message
def find_item_notes(soup):
	notes_div = soup.find(id = "details-notes")
	if notes_div:
		notes = notes_div["td"].text
		print(notes)
		return notes
	else:
		print("no notes")
		return False

#creates request object and parses w/ BeautifulSoup
def make_soup(url):
	resp = requests.get(url)
	resp = resp.content
	soup_contents = BeautifulSoup(resp, 'html.parser')
	return soup_contents


## makes search request  looks for subjects
def make_request(primary_url, error_cell_number):
	print("Making request for new data using primary_url")
	soup = make_soup(primary_url)

	## if WorldCat brings up error that no results found with primary_url, then try secondary_url
	if soup.find_all(class_ = "error-results", id = "div-results-none"):
		error_message = "Error with primary URL"
		write_to_excel(error_cell_number, error_message)
		print(error_message)
		request_without_error = False

	## no error with primary_URL, now find out if on menu page and look for subjects 
	else:
		request_without_error = True
		##need to find out if on menu page or specific item page
		baseurl = "https://www.worldcat.org"
		
		## follow the first href on menu page and get its soup
		menu = soup.find(class_ = "menuElem")
		menu_items = menu.find_all(class_= "result details")
		print("The menu items exist")


		## verify in the correct language in the first link
		verifying_language = menu_items[0].find(class_= "itemLanguage").text
		if verifying_language == "Judeo-Arabic" or "Hebrew":
			href = menu_items[0].find("a")['href']
			complete_url = baseurl + href
			print("Making request for new data using HREF from 1st in menuElem")
			item_page_soup = make_soup(complete_url)
			valid_entries = True
		## check out the language in the second link, if no second link, the except 
		elif: 
			try:			
				verifying_language = menu_items[1].find(class_= "itemLanguage").text
				if verifying_language == "Judeo-Arabic" or "Hebrew":
					href = menu_items[1].find("a")['href']
					complete_url = baseurl + href
					print("Making request for new data using HREF from 2nd in menuElem")
					item_page_soup = make_soup(complete_url)
					valid_entries = True
			# meaning, there is no second link on menu
			except:
				continue
		## neither of the first two entries in judeo arabic or hebrew, or there is only one entry and it isn't in JA or H:
		else:
			valid_entries = False
			error_message = "Wrong language"
			write_to_excel(error_cell_number, error_message)
			print(error_message)			
			return valid_entries   ### --> in this case, only thing returned is false.


		## able to find a valid entry, will look for and return the subject string if exists, if not, returns False
		if valid_entries == True:
			subject_string_or_false = find_subjects(item_page_soup, error_cell_number)
			item_description_or_false = find_page_number_description(item_page_soup)
			item_notes_or_false = find_item_notes(item_page_soup)


		### need to find number of pages, and other notes, if possible, and return a tuple 

		print(subject_string_or_false)
		return subject_string_or_false


	# #returns false if not able to reach an appropriate URL
	# else:
	# 	return request_without_error






##tries to make request with two different URLS and looks for subjects
# def make_request(primary_url, secondary_url):
# 	print("Making request for new data using primary_url")
# 	soup = make_soup(primary_url)

# 	## if WorldCat brings up error that no results found with primary_url, then try secondary_url
# 	if soup.find_all(class_ = "error-results", id = "div-results-none"):
# 		print("found an error with the first url")

# 		### if there is a secondary url to try, try it. 
# 		try: 
# 			print("Making request for new data using SECONDARY Url")
# 			soup = make_soup(secondary_url)
# 			if soup.find_all(class_ = "error-results", id = "div-results-none"):
# 				print("Found an error with the first and second url!")
# 				##  secondary_url returned error, as well
# 				request_without_error = False

# 			## secondary_url worked!	
# 			else:
# 				request_without_error = True

# 		### if there is no secondary url, out of luck, ultimately false
# 		except:
# 			request_without_error = False

# 	# meaning the primary url worked fine		
# 	else:
# 		request_without_error = True

# 	## no error with URLs, now find out if on menu page and look for subjects 
# 	if request_without_error == True:

# 		##need to find out if on menu page or specific item page
# 		the_menu_exists = soup.find(class_ = "menuElem")
# 		print("the_menu_exists")

# 		## if we are indeed on the menu page... 			
# 		## follow the first href and get its soup
# 		if the_menu_exists:
# 			baseurl = "https://www.worldcat.org"
# 			menu_items = the_menu_exists.find_all(class_= "result details")
# 			print("The menu items exist")

# 			## some of the titles are the same as Arabic works
# 			verifying_language = menu_items[0].find(class_= "itemLanguage").text
# 			if verifying_language == "Judeo-Arabic" or "Hebrew":
# 				href = menu_items[0].find("a")['href']
# 				complete_url = baseurl + href
# 				print("Making request for new data using HREF from menuElem")
# 				soup = make_soup(complete_url)
# 			else:
# 				print("Got mixed up with an Arabic text")


# 		#if not on menu page, try assuming we are on the item page. find_subjects will return false otherwise.

# 		## will return the subject string if exists, if not, returns False
# 		subject_string_or_false = find_subjects(soup)

# 		### need to find number of pages, and other notes, if possible, and return a tuple 

# 		print(subject_string_or_false)
# 		return subject_string_or_false
# 	#returns false if not able to reach an appropriate URL
# 	else:
# 		return request_without_error

def iterate_excel_file():
	wb = load_workbook(filename = FNAME, read_only = True)
	sheet = wb['Judeo-Arabic']
	sheet = wb.active   ## got this from https://stackoverflow.com/questions/49159245/python-error-on-get-sheet-by-name

	# current_number = 2  ## to start with ----> CHANGE BACK!!!!

	##it happened around 90, 114, 139
	current_number = 289


	count_times_written = 0
	count_times_NOT_written = 0 
	title_cell_number = "H" + str(current_number)
	author_cell_number = "G" + str(current_number)

	subject_cell_number = "W" + str(current_number)
	error_cell_number= "AD" + str(current_number)


	#for all the rows of items
	while sheet[title_cell_number].value != None:


	# for i in range(3):
		#if item does not already have subjects assined to it
	#		print("value of subject cell is None!!!")
		print("Title cell number:")
		print(title_cell_number)
		current_title = sheet[title_cell_number].value


		### making sure the title is more than one word long
		if len(current_title) > 1:
			print(current_title)
			current_author = sheet[author_cell_number].value
			print(current_author)
			##pass the title and author into assemble_url, which makes it into a url can make request with
			if current_author != "":
				assembled_url = assemble_url(current_title, current_author)
				print(assembled_url)
			else:
				assembled_url = assembled_url(current_title)
	## could also make it so that it always puts in the the author, if putting in a "" author doesn't mess it up. 
			##making the request with assembled URL, if there's an error, will return false
			result = make_request(assembled_url, error_cell_number) 

			## meaning there was a valid entry 
			# if result != False:

			## meaning there was a valid entry and it went forth to check subjects, notes, description
			if len(result) >1:



				subjects = result[0]
				description = result[1]
				notes = result[2]


				#it was able to find subjects
				if subjects != False:
					write_to_excel(subject_cell_number, subjects)
					count_times_written +=1 

				if description != False:



				else:
					count_times_NOT_written += 1

		## writing in excel saying it was skipped because of non-distinct name
		else:
			error_message = "Indistinct title"
			write_to_excel(error_cell_number, error_message)





		print("\n###################\n")
		current_number += 1
		title_cell_number = "H" + str(current_number)

		#there is already a subject assigned to the cell number 
		# else:
		# 	continue

	print("Times written = ")
	print(count_times_written)
	print("Times NOT written = ")
	print(count_times_NOT_written)
	print("count_times_recorded_error_note:")
	print(count_times_recorded_error_note)

	#return sheet




######### To demo a part of the script #######
# example_assembled = 'https://www.worldcat.org/title/kohelet-im-sharh-ha-arvi-ha-meduberet-ben-ha-am-ve-im-perush-shema-shelomoh'

# example_assembled = "https://www.worldcat.org/title/-20160603/oclc/6914317075&referer=brief_results"
# made = make_request(example_assembled, None)
# print(made)


######### TO RUN THE SCRIPT, UNCOMMENT BELOW! ######
iterate_excel_file()



######  SCRIPT TO-DO LIST #####
# if starts with al, make a contingency url, that eliminates it DONEZO
# if not, choose the second link  DONEZO
# run the loop again, but if if the value of the subject_tags column is none, then get the info DONEZO
# if lands on a page with many items, choose the first one, look for subjects DONEZO

## possibly investigate second href on the menu page, as well

## improve what happens with tags with commas in them like:
#"Esther,; Queen; of; Persia; Bible; stories,; Judeo; Arabic; Esther; Old; Testament"
# or " Judeo-Arabic; literature; Jews; History; To; 70; AD"


### make a contingency before getting something from the menu that it is in hebrew or judeo-arabic

## need to make it so file format doesn't die after writing on it.


#instead of constructing url of title, make a url of title and author, if available 
#also, confirm the text is in hebrew / j-a


### add a flagging to the spread sheet in a different column to say there was an error, couldn't find


