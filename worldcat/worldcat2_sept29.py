import json
import requests
import csv
import secrets as secrets
from bs4 import BeautifulSoup
import pandas as pd
import re

from openpyxl import Workbook 
from openpyxl import load_workbook
import openpyxl

# api_key = secrets.api_key

# FNAME = "Judeo_arabic_spreadsheet copy.csv"
FNAME = "Judeo_arabic_spreadsheet copy.xlsx"

alternative_place_spellings_dict= {"Alexandria": ("Alexandria", "aleksandria", "aleksandrya", "Aleksandriya", "Eskendereyya", "al-ʾIskandariyya", "Rakote", 7), 
"Aden": ("Aden", "Adin", 2),
"Aleppo": ("Aleppo", 'Halab', "Halabi", "Halaba", 4),
"Algiers": ("Algiers","Alger", 2),
"Baghdad": ("Baghdad", "Bagdad", 2),
"Bombay": ("bombay", 1),
"Calcutta": ("Calcutta", "Kolkata", 2),
"Constantine": ("Constantine", "Qacentina", "Kasantina", 3),
"Djerba": ("Djerba", "gerba", "gerbah", "jerbah", "bjerba", "bjerbah", 7),
"Jerusalem": ("Jerusalem", "yerushalayim", 2),
"Livorno": ("Livorno", "Leghorn", 2),
"Oran": ("Oran", "Wehran", 2),
"Sousse": ("Sousse", "Susah", 2),
"Tunis": ("tunis", 1)
}


## creating a class for each publication
class Publication(object):
    def __init__(self, title):
        self.title = title
        self.subject_tags = ""
    def __str__(self):
        return subject_tags
    pass

#opens excel file, writes subjects to the given subject_tag cell number --> WILL NEED TO CHANGE SPECIFIED CELL # LATER
def write_to_excel(write_to_cell_number, content_to_write):
    wb = Workbook()
    xfile = openpyxl.load_workbook(filename = FNAME, read_only=False, keep_vba=True)
     # prior to sept 28
    sheet = xfile.get_sheet_by_name('Judeo-Arabic')

    #print(sheet[write_to_cell_number].value)
    sheet[write_to_cell_number].value = content_to_write
    xfile.save(FNAME)


## perform reg ex on a term that is going to be used in the URL query
def reg_ex(query_term):
    complete_query_string = ''
    improved_query_term = query_term.replace('[', ' ').replace(']', ' ').replace('.', ' ').replace(':', ' ').replace("ʻ", "").replace("(", "").replace(")", "").replace("ʼ", "").replace("'", "").replace(",", "").lower().split()
    for word in improved_query_term:
        complete_query_string += word + "-"
    complete_query_string = complete_query_string[:-1].replace("--", '-')

    return complete_query_string

#creates request object and parses w/ BeautifulSoup
def make_soup(url):
    resp = requests.get(url)
    resp = resp.content
    soup_contents = BeautifulSoup(resp, 'html.parser')
    return soup_contents


#takes give title, eliminates uncessary characters and tokenizes, creates primary url. 

#if first word in title is "al" creates, secondary url
def assemble_url(a_title, current_place_lst):
    title= reg_ex(a_title)
    baseurl = "https://www.worldcat.org/search?q=" + title
    print("LENTH OF A PLACE LIST: ", len(current_place_lst))
    if len(current_place_lst) > 0:
        print("Greater than 0")
        # if len(current_place_lst) > 1:
        url_lst = [baseurl + "+" + reg_ex(current_place_lst[i]) for i in range(current_place_lst[-1])]
        print("this is the url list!!!", url_lst)
        # else:
        #     print("length is 1")
        #     url_lst = [baseurl + "+" + reg_ex(current_place_lst[0])]
        # url_lst = []
        # for i in range(:len(a_place_tuple))
        #     print("this is i!!!! ", a_place_tuple[i])
        #     place = reg_ex(a_place_tuple[1])
        #     primary_url = baseurl + "-" + place
        #     print(primary_url)
    else:
        # print("COULD NOT FIND A PLACE. ITS AN EMPTY TUPLE")
        url_lst = [baseurl]
        print(url_lst)
# url's are now a lst of urls. will return a lst either way
    return url_lst
#
##try to find subject div, if doesn't exist, meaning there are no subjects, return false and record the error in error_cell_number
##if there are subjects, return subject string. 
def find_subjects(soup, error_cell_number):
    #if the item has subject terms
    if soup.find_all(id = "subject-terms"):
        subject_div = soup.find_all(id = "subject-terms")
        # subject_div = subject_div.find_all('li', attrs={'class' :"subject-term"})
        subject_list = []
        subject_string = ""

        for line in subject_div:
            subjects = line.text.replace(".", "").split("\n")
            for a_line in subjects:
                a_line = a_line.split(" -- ")


        ## finally i've got a line that is all clean
                for word in a_line:
                    ## if word isn't empty ''
                    if len(word) >0:
                    ## getting rid of unnecessary comma cause i'm going to join them with semi-colon
                        if word[-1] == ",":
                            new_word = word[:-1] 
                        else:
                            new_word = word
                        #now, i have the official "new word" either way
                        if new_word not in subject_list:
                            subject_list.append(new_word)

        subject_string = "; ".join(subject_list)
        return subject_string
    #the item doesn't have subject terms 
    else:
        ##couldn't get subjects 
        print("Could not find subjects")
        error_message = "No subjects"
        write_to_excel(error_cell_number, error_message)
        return False            


## finds and returns description or false. does not record an error message
def find_page_number_description(soup):
    try:
        # description_div = soup.find(id = "details-description")
        # description = description_div.find("td").text
        # print(description)
        description = soup.find("tr", id = 'details-description').find("td").text
        print(description)
        return description
    except:
        print("no description")
        return False

## finds and returns notes or false. does not record an error message
def find_item_notes(soup):
    # if notes_div exists:
    try:
        # notes_div = soup.find(id = "details-notes")
        # notes = notes_div.find("td")
        # notes = str(notes).replace("<br/>", " ").replace("<td>", "").replace("</td>", "")
        # print(notes)
        notes = soup.find("tr", id = 'details-notes').find("td").text
        print(notes)
        return notes
    # no notes_div exists
    except:
        print("no notes")
        return False

## finds and returns OCLC number or false. does not record an error message
def find_item_OCLC(soup):
    # if notes_div exists:
    # try:
        # OCLC_div = soup.find(id = "details-oclcno")
        # OCLC = notes_div.find("td")
        # OCLC = str(OCLC).replace("<br/>", " ").replace("<td>", "").replace("</td>", "")
        # print(OCLC)
    OCLC = soup.find("tr", id = 'details-oclcno').find("td").text
    print(OCLC)
    return OCLC
    # no notes_div exists
    # except:
    #     print("no OCLC")
    #     return False

def find_item_genre(soup):
    # if genre_div exists:
    try:
        # genre_div = soup.find(id = "details-genre")
        # print(genre_div)
        # genre = notes_div.find("td")
        # print(genre)
        # genre = str(genre).replace("<br/>", " ").replace("<td>", "").replace("</td>", "")
        # print(genre)
        genre = soup.find("tr", id = 'details-genre').find("td").text
        print(genre)
        return genre
    # no notes_div exists
    except:
        print("no genre")
        return False

def find_item_provenance(soup):
    # if genre_div exists:
    try:
        # provenance_div = soup.find(id = "details-provenance")
        # provenance = provenance_div.find("td")
        # provenance = str(provenance).replace("<br/>", " ").replace("<td>", "").replace("</td>", "")
        # print(provenance)
        provenance = soup.find("tr", id ='details-provenance').find("td").text
        print(provenance)
        return provenance
    # no notes_div exists
    except:
        print("no provenance")
        return False

## makes search request looks for subjects
def make_request(url_lst, error_cell_number):
    print("Making request for new data using url_lst")
    legit_url = False
    while legit_url == False:
        for i in url_lst:
            soup = make_soup(i)
            print("got to line 233")
            #check if error with primary url
            if soup.find_all(class_ = "error-results", id = "div-results-none"):
                print("error with primary url")
                error_message = "Error with primary URL"
                # write_to_excel(error_cell_number, error_message)
                print(error_message)
                request_without_error = False

                #go to next iterator in the URL list 
                legit_url = False
                print("\n\n\nNEEDS TO CONTINUE HERE\n\n\n")
                continue
                # return request_without_error

            ## no error with primary_URL, now find out if on menu page and look for subjects 
            else:
                print("no error with primary_URL")
                request_without_error = True

                ##need to find out if on menu page or specific item page
                baseurl = "https://www.worldcat.org"
                
                ## follow the first href on menu page and get its soup
                menu = soup.find(class_ = "menuElem")
                menu_items = menu.find_all(class_= "result details")
                print("The menu items exist")

                ## verify in the correct language in the first link
                verifying_language = menu_items[0].find(class_= "itemLanguage").text
                if verifying_language == "Judeo-Arabic" or "Hebrew":
                    href = menu_items[0].find("a")['href']
                    complete_url = baseurl + href
                    print("Making request for new data using HREF from 1st in menuElem, which is:")
                    print(complete_url)
                    item_page_soup = make_soup(complete_url)
                    valid_entries = True

                ## check out the language in the second link, if no second link, the except 
                elif menu_items[1].find(class_= "itemLanguage").text == "Judeo-Arabic" or "Hebrew":
                    try:            
                        verifying_language = menu_items[1].find(class_= "itemLanguage").text
                        if verifying_language == "Judeo-Arabic" or "Hebrew":
                            href = menu_items[1].find("a")['href']
                            complete_url = baseurl + href
                            print("Making request for new data using HREF from 2nd in menuElem")
                            item_page_soup = make_soup(complete_url)
                            valid_entries = True
                    # meaning, there is no second link on menu
                    except:
                        print("got to except. neither of first two results valid")
                        #go to the next iterator 
                        legit_url = False
                        valid_entries = False

                ## neither of the first two entries in judeo arabic or hebrew, or there is only one entry and it isn't in JA or H:
                else:
                    valid_entries = False
                    legit_url = False
                    error_message = "Wrong language"
                    # write_to_excel(error_cell_number, error_message)
                    print(error_message)            
                    return valid_entries   ### --> in this case, only thing returned is false.
        #end of iterator
    #end of while statement


# else:
#     pass

        ## able to find a valid entry, will look for and return the subject string if exists, if not, returns False
        if valid_entries == True:
            subject_string_or_false = find_subjects(item_page_soup, error_cell_number)
            item_description_or_false = find_page_number_description(item_page_soup)
            item_notes_or_false = find_item_notes(item_page_soup)
            item_OCLC_or_false = find_item_OCLC(item_page_soup)
            item_genre_or_false = find_item_genre(item_page_soup)
            item_provenance_or_false= find_item_provenance(item_page_soup)

        ### need to find number of pages, and other notes, if possible, and return a tuple 
        return_tuple = (subject_string_or_false, item_description_or_false, item_notes_or_false, item_OCLC_or_false, item_genre_or_false, item_provenance_or_false)
        print(len(return_tuple))
        print(return_tuple)
        return return_tuple



def iterate_excel_file():

    # prior to sept 28, this is how you loaded data. i now cant remember why i did it this way
    wb = load_workbook(filename = FNAME, read_only = True)
    ## getting a specific sheet from the XL
    sheet = wb['Judeo-Arabic']
    sheet = wb.active   ## got this from https://stackoverflow.com/questions/49159245/python-error-on-get-sheet-by-name
    

# post sept 28, i thought i could put it in csv format instead
    # with open(FNAME, 'rb') as csvfile:
    #     sheet = csv.reader(csvfile)

    ## Cell row to start with 
    current_number = 2  ############ ---> Original start row = 2


    current_number = 66

    title_cell_number = "H" + str(current_number)
    author_cell_number = "G" + str(current_number)
    place_cell_number = "L" + str(current_number)
    genre_cell_number = "W" + str(current_number)
    subject_cell_number = "X" + str(current_number)
    description_cell_number = "Y" + str(current_number)
    notes_cell_number = "Z" + str(current_number)
    error_cell_number= "AA" + str(current_number)
    oclc_cell_number = "AH" + str(current_number)
    provenance_cell_number = "AI" + str(current_number)
    #for all the rows of items.

    #this used to be like this prior to sept 28
    # while sheet[title_cell_number].value != None:

    #just trying this out on sept 30 to see if i can isolate the problem
    while current_number < 68:
    # for i in range(3):
        #if item does not already have subjects assined to it
    #       print("value of subject cell is None!!!")
        print("Title cell number:")
        print(title_cell_number)
        current_title = sheet[title_cell_number].value

        ### making sure the title is more than one word long
        if len(current_title) > 1:
            # print(current_title)
            #current place now needs to map to a dictionary of alternative place spellings
            print("THIS SI THE CURRENT PLACE!")
            current_place = sheet[place_cell_number].value
            # print(current_place)

            if current_place in alternative_place_spellings_dict:
                # print((alternative_place_spellings_dict.items()))
                if alternative_place_spellings_dict[current_place][-1] != 1:
                    current_place_lst= list(alternative_place_spellings_dict[current_place])
                else:
                    current_place_lst = list(alternative_place_spellings_dict[current_place])

                    # print("CURRENT PLACE LIST: ", current_place_lst)
                    # print(current_place_lst)
                    # print("WENT TO THE DICT !!!!!!!!!!!!!!!!!!!!")
            else:
                current_place_lst = []

                # print("DIDNT GO TO THE DICT!!!")

            ##pass the title and current_place_lst into assemble_url(), which makes it into a url can make request with
            
            # if current_author != "":
            assembled_url = assemble_url(current_title, current_place_lst)
            #     print(assembled_url)
            # else:
            #     assembled_url = assembled_url(current_title)

            #                     assembled_url = assembled_url(current_title)

    ## could also make it so that it always puts in the the author, if putting in a "" author doesn't mess it up. 
            ##making the request with assembled URL, if there's an error, will return false
            result = make_request(assembled_url, error_cell_number) 

            ## meaning there was a valid entry 
            # if result != False:

            # <1 would mean was no valid entry and therefore, didn't get subjects
            ## meaning there was a valid entry and it went forth to check subjects, notes, description
            if result != False:
            # if len(result) >1:
                subjects = result[0]
                description = result[1]
                notes = result[2]
                OCLC = result[3]
                genre = result[4]
                provenance = result[5]


                #it was able to find subjects
                if subjects != False:
                    write_to_excel(subject_cell_number, subjects)

                #if found a description
                if description != False:
                    write_to_excel(description_cell_number, description)

                #if found notes
                if notes != False:
                    write_to_excel(notes_cell_number, notes)

                #if found OCLC Number
                if OCLC != False:
                    write_to_excel(oclc_cell_number, OCLC)

                #if found genre 
                if genre != False:
                    write_to_excel(genre_cell_number, genre)

                #if found OCLC provenance
                if provenance != False:
                    write_to_excel(provenance_cell_number, provenance)



        ## writing in excel saying it was skipped because of non-distinct name
        else:
            error_message = "Indistinct title"
            write_to_excel(error_cell_number, error_message)


        print("\n###################\n")
        current_number += 1

        ## increase the cell numbers by 1
        title_cell_number = "H" + str(current_number)
        author_cell_number = "G" + str(current_number)
        place_cell_number = "L" + str(current_number)
        genre_cell_number = "W" + str(current_number)
        subject_cell_number = "X" + str(current_number)
        description_cell_number = "Y" + str(current_number)
        notes_cell_number = "Z" + str(current_number)
        error_cell_number= "AA" + str(current_number)
        oclc_cell_number = "AH" + str(current_number)
        provenance_cell_number = "AI" + str(current_number)




    #return sheet




######### To demo a part of the script #######
# example_assembled = 'https://www.worldcat.org/title/kohelet-im-sharh-ha-arvi-ha-meduberet-ben-ha-am-ve-im-perush-shema-shelomoh'

# example_assembled = "https://www.worldcat.org/title/-20160603/oclc/6914317075&referer=brief_results"
# made = make_request(example_assembled, None)
# print(made)


######### TO RUN THE SCRIPT, UNCOMMENT BELOW! ######
iterate_excel_file()



######  SCRIPT TO-DO LIST #####
# if starts with al, make a contingency url, that eliminates it DONEZO
# if not, choose the second link  DONEZO
# run the loop again, but if if the value of the subject_tags column is none, then get the info DONEZO
# if lands on a page with many items, choose the first one, look for subjects DONEZO

## possibly investigate second href on the menu page, as well

## improve what happens with tags with commas in them like:
#"Esther,; Queen; of; Persia; Bible; stories,; Judeo; Arabic; Esther; Old; Testament"
# or " Judeo-Arabic; literature; Jews; History; To; 70; AD"


### make a contingency before getting something from the menu that it is in hebrew or judeo-arabic

## need to make it so file format doesn't die after writing on it.


#instead of constructing url of title, make a url of title and author, if available 
#also, confirm the text is in hebrew / j-a


### add a flagging to the spread sheet in a different column to say there was an error, couldn't find


